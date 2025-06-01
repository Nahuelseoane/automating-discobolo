import pandas as pd
from datetime import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config.config import TRANSFER_FILE, BANK_FILE, MONTH


# Get current month index (1 = January, 2 = February, ...)
current_month_number = datetime.now().month

month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
current_month_sheet = month_names[current_month_number - 1]
print(f"   📂 Updating sheet: {current_month_sheet}")

# Load bank file
df_bank = pd.read_excel(BANK_FILE, skiprows=1, engine="openpyxl")

# Delete "descripción" column (duplicate one)
df_bank = df_bank.drop(columns=["Descripción"])

# ✅ Fix column names for consistency
df_bank.columns = df_bank.columns.str.strip()
df_bank = df_bank.rename(columns={
    "Número Secuencia": "N° Secuencia",
    "Descripción Extendida": "Descripción"
})

# ✅ Ensure column order matches the master file
expected_columns = ["N° Secuencia", "Fecha", "Descripción", "Importe", "Saldo"]
df_bank = df_bank[expected_columns]

# ✅ Load the master file without modifying formatting
wb = load_workbook(TRANSFER_FILE)
last_sequence = None
if current_month_sheet not in wb.sheetnames:
    print(f"❌ Error: Sheet '{current_month_sheet}' not found in master file.")

    wb.create_sheet(MONTH)
    wb.save(TRANSFER_FILE)
    print(f"✅ {MONTH} sheet created successfully in {TRANSFER_FILE}")

previous_sheet = wb[month_names[current_month_number - 2]]
current_sheet = wb[current_month_sheet]
# Check if current sheet is empty (beyond headers)
is_empty = all(
    all(cell.value is None for cell in row)
    for row in current_sheet.iter_rows(min_row=2, values_only=False)
)
# Sequence number
if is_empty:
    header_row = list(previous_sheet.iter_rows(
        min_row=1, max_row=1, values_only=False))[0]
    # Copy first row from previous sheet
    for col_idx, source_cell in enumerate(header_row, start=1):
        target_cell = current_sheet.cell(row=1, column=col_idx)
        target_cell.value = source_cell.value
        # Copy style
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    # Copy column widths
    for col_idx in range(1, previous_sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in previous_sheet.column_dimensions:
            current_sheet.column_dimensions[col_letter].width = previous_sheet.column_dimensions[col_letter].width
# Copy Sequence number
    last_sequence = int(previous_sheet.cell(row=2, column=1).value)


# ✅ Find the last recorded sequence in master file
if last_sequence == None:
    seq_column = None
    for col in current_sheet.iter_cols(min_row=1, max_row=1, values_only=True):
        if "N° Secuencia" in col:
            seq_column = col.index("N° Secuencia") + 1
            break

    if not seq_column:
        print("❌ Could not find 'N° Secuencia' column in master file.")
        exit()

    # Seq number
    for row in current_sheet.iter_rows(min_row=2, values_only=True):
        if row[seq_column - 1]:  # Ensure it's not empty
            last_sequence = int(row[seq_column - 1])
            break

    if last_sequence is None:
        print("❌ Could not determine last sequence number in master file.")
        exit()
print(f"   📌 Last recorded sequence in master: {last_sequence}")

# ✅ Filter only new transactions
df_new = df_bank[df_bank["N° Secuencia"] > last_sequence]
df_new = df_new.sort_values(by="N° Secuencia", ascending=True)  # Correct order

if df_new.empty:
    print("   ✅ No new transactions to add.")
    exit()

print(f"📌 Adding {len(df_new)} new transactions to {current_month_sheet}.")

# ✅ Insert new rows before existing transactions (to preserve format)
for row in df_new.itertuples(index=False):
    # Always insert at row 2 (below column headers)
    current_sheet.insert_rows(2)
    for col_num, value in enumerate(row, start=1):
        current_sheet.cell(row=2, column=col_num, value=value)

# ✅ Save without losing formatting
wb.save(TRANSFER_FILE)
wb.close()
print("   ✅ Transfer file updated successfully")
